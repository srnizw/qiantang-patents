import openpyxl
import json
import os
import re

BASE = os.path.dirname(os.path.abspath(__file__))
ZHUANLI = os.path.join(BASE, 'zhuanliVX')

# ---- Classification: 车药芯化智 ----
# 车=智能车 药=生物药 芯=中国芯 化=新化材 智=智造端(AI驱动)
CATEGORY_KEYWORDS = {
    '车': [
        '汽车', '车辆', '驾驶', '交通', '发动机', '新能源车', '充电桩', '动力电池',
        '自动驾驶', '无人驾驶', '智能座舱', '车联网', '底盘', '轮胎', '变速箱',
        '电动车', '混合动力', '车载', '行车', '泊车', '车道', '路况', '交通安全',
        '尾气', '内燃机', '转向', '制动', '悬挂', '气囊', '安全带', '差速器',
        '轮毂', '雨刮', '离合器', '曲轴', '凸轮轴', '涡轮增压',
    ],
    '药': [
        '药物', '制剂', '疫苗', '抗体', '病理', '药理', '中药', '生物医药',
        '医疗器械', '手术', '靶向', '抗菌', '抑菌', '杀菌', '肿瘤', '抗癌',
        '免疫', '代谢', '肽', '酶', '多糖', '提取物', '药用', '药理学',
        '医学', '给药', '剂型', '治疗', '临床', '诊断', '基因', '细胞', '蛋白',
        '毒理', '药代', '药动', '药用植物', '活性成分', '有效部位',
    ],
    '芯': [
        '芯片', '半导体', '集成电路', '处理器', '射频', '微电子', '嵌入式',
        'FPGA', '晶圆', 'MEMS', '光电子', '毫米波', '硅基', '氮化镓', '碳化硅',
        '存储器', '模数转换', '锁相环', '振荡器', '放大器', '印制电路板',
        '电路设计', '模数变换', 'ADC', 'DAC', '滤波器', '天线', '封装基板',
        'EDA', '集成电路设计', 'SoC', 'ASIC',
    ],
    '化': [
        '化工', '化学', '催化', '涂料', '染料', '试剂', '萃取', '电解',
        '高分子', '石墨烯', '碳纤维', '膜分离', '聚合', '树脂', '橡胶', '塑料',
        '表面活性剂', '分散剂', '阻燃', '防腐', '复合材料', '功能材料',
        '光催化', '电催化', '吸附', '脱硫', '脱硝', '纺丝', '纺纱', '印染',
        '着色', '颜料', '有机合成', '无机材料', '陶瓷', '合金', '涂层', '镀层',
        '冶金', '水泥', '玻璃', '制备', '分离纯化', '结晶', '蒸馏',
        '纳米材料', '纳米管', '纳米线', '纳米颗粒', '纳米片',
        '提铜', '提锌', '提金', '提银', '浸出', '冶炼', '湿法', '火法',
        '废物', '废水', '废气', '回收', '提纯', '萃取剂', '离子交换',
        '氧化反应', '还原反应', '降解', '发泡', '乳化', '固化',
        '金属材料', '无机非金属', '生物质', '生物基材料',
    ],
    '智': [
        '人工智能', '深度学习', '神经网络', '机器学习', '计算机视觉',
        '自然语言处理', '大语言模型', '具身智能', '智能机器人', '人形机器人',
        '数字孪生', '工业互联网', '物联网', '边缘计算', '云计算',
        '智能制造', '智能检测', '智能识别', '智能控制', '智能决策',
        '自动化生产线', '机器视觉', '3D打印', '激光加工', '数控机床',
        '故障诊断', '预测维护', '质量控制', '精密加工', '无损检测',
        '信息安全', '加密', '区块链', '数据融合', '数据采集',
        '目标检测', '语义分割', '图像处理', '模式识别', '知识图谱',
        '自主导航', '无人车', '无人机', '无人船',
    ],
}

# Check order: more specific categories first, generic ones last
CAT_ORDER = ['车', '药', '芯', '智', '化']


def classify(name, abstract):
    text = (name + ' ' + abstract)

    # Pass 1: count exact keyword matches
    scores = {}
    for cat in CAT_ORDER:
        score = 0
        for kw in CATEGORY_KEYWORDS[cat]:
            if kw in text:
                score += 1
        scores[cat] = score

    best = max(scores, key=scores.get)
    if scores[best] > 0:
        return (best, scores[best])

    # Pass 2: broader fallback — looser keyword matching, lower confidence
    broad = {
        '车': ['车', '驾驶', '制动', '底盘'],
        '药': ['药', '医', '患者', '肿瘤', '细胞', '病毒', '菌株'],
        '芯': ['芯片', '电路', '晶体管', '信号处理', '微处理器'],
        '智': ['算法', '识别', '预测', '学习', '网络模型', '优化', '定位', '导航', '控制', '通信'],
        '化': ['材料', '化学', '合成', '制备', '纤维', '金属', '矿物', '提取', '分离', '降解', '废水', '催化'],
    }
    broad_scores = {}
    for cat in CAT_ORDER:
        s = 0
        for kw in broad[cat]:
            if kw in text:
                s += 1
        broad_scores[cat] = s
    best2 = max(broad_scores, key=broad_scores.get)
    if broad_scores[best2] > 0:
        return (best2, broad_scores[best2] * 0.5)  # lower confidence

    # Pass 3: character-level heuristic, lowest confidence
    if any(c in text for c in ['药', '医']):
        return ('药', 0.1)
    if any(c in text for c in ['车', '轮', '驾']):
        return ('车', 0.1)
    if any(c in text for c in ['芯片', '集成电路', '半导体']):
        return ('芯', 0.1)
    if any(c in text for c in ['材料', '化学', '催化', '纳米']):
        return ('化', 0.1)
    if any(c in text for c in ['算法', '识别', '学习', '智能']):
        return ('智', 0.1)

    return ('化', 0)


def clean(s):
    if s is None:
        return ''
    s = str(s).strip()
    # remove excessive whitespace
    s = re.sub(r'\s+', ' ', s)
    return s


def read_all_sheets(filepath):
    """Read all rows from all sheets, return list of (name, unit, abstract)."""
    results = []
    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sn in wb.sheetnames:
        ws = wb[sn]
        if ws.max_row < 2:
            continue

        # Detect column mapping from first row headers
        headers = [clean(ws.cell(row=1, column=c + 1).value) for c in range(ws.max_column)]

        # Find indices for name, unit, abstract
        name_cols = [i for i, h in enumerate(headers)
                     if any(kw in h for kw in ('专利名称', '发明名称', '专利标题'))
                     or any(h == kw for kw in ('标题 (中文)', '标题'))]
        if not name_cols:
            name_cols = [i for i, h in enumerate(headers) if '名称' in h or '标题' in h]

        unit_cols = [i for i, h in enumerate(headers)
                     if any(kw in h for kw in ('申请人', '专利权人', '单位', '第一申请人'))
                     or ('申请' in h and ('人' in h or '者' in h))
                     or ('专利权' in h and '人' in h)]

        abstract_cols = [i for i, h in enumerate(headers) if '摘要' in h]

        if not name_cols:
            continue

        name_idx = name_cols[0]
        unit_idx = unit_cols[0] if unit_cols else None
        abstract_idx = abstract_cols[0] if abstract_cols else None

        for row_idx in range(2, ws.max_row + 1):
            name = clean(ws.cell(row=row_idx, column=name_idx + 1).value)
            if not name:
                continue
            unit = clean(ws.cell(row=row_idx, column=unit_idx + 1).value) if unit_idx is not None else ''
            abstract = clean(ws.cell(row=row_idx, column=abstract_idx + 1).value) if abstract_idx is not None else ''

            # Skip header-like rows that accidentally match
            if name in ('专利名称', '发明名称', '名称') or len(name) < 2:
                continue

            category, score = classify(name, abstract)

            results.append({
                'n': name,
                'u': unit,
                'a': abstract,
                'c': category,
                's': score,
                '_src': os.path.basename(filepath),
            })

    wb.close()
    return results


def main():
    xlsx_files = sorted([f for f in os.listdir(ZHUANLI) if f.endswith('.xlsx')])

    all_patents = []
    stats = {}

    for f in xlsx_files:
        fpath = os.path.join(ZHUANLI, f)
        print(f'Reading: {f} ...', end=' ', flush=True)
        patents = read_all_sheets(fpath)
        all_patents.extend(patents)
        stats[f] = len(patents)
        print(f'{len(patents)} patents')

    print(f'\nTotal: {len(all_patents)} patents')

    # Unit overrides (fix mislabeled data at source)
    UNIT_OVERRIDES = {
        '授权有效专利2025.xlsx': '东南大学',
    }
    for p in all_patents:
        src = p.pop('_src')
        if src in UNIT_OVERRIDES:
            p['u'] = UNIT_OVERRIDES[src]

    # Category distribution
    cat_counts = {}
    for p in all_patents:
        cat_counts[p['c']] = cat_counts.get(p['c'], 0) + 1
    print('Category distribution:', cat_counts)

    # Deduplicate by name (keep first occurrence)
    seen = set()
    deduped = []
    dup_count = 0
    for p in all_patents:
        key = p['n']
        if key in seen:
            dup_count += 1
            continue
        seen.add(key)
        deduped.append(p)
    if dup_count:
        print(f'Removed {dup_count} duplicates')

    # Sort: relevance score (desc) → university priority → name
    UNI_PRIORITY = [
        '浙江工商大学', '杭州电子科技大学', '浙江理工大学',
        '浙江水利水电学院', '浙江大学', '东南大学',
    ]

    def sort_key(p):
        u = p['u']
        uni_rank = len(UNI_PRIORITY)
        for i, uni in enumerate(UNI_PRIORITY):
            if uni in u:
                uni_rank = i
                break
        return (-p['s'], uni_rank, p['n'])

    deduped.sort(key=sort_key)

    # Output JSON
    out_path = os.path.join(BASE, 'patents.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(deduped, f, ensure_ascii=False, separators=(',', ':'))

    # Output JS (for file:// compatibility)
    js_path = os.path.join(BASE, 'patents.js')
    with open(js_path, 'w', encoding='utf-8') as f:
        f.write('window.PATENT_DATA=')
        json.dump(deduped, f, ensure_ascii=False, separators=(',', ':'))
        f.write(';')

    print(f'\nWritten {len(deduped)} patents to patents.json and patents.js')

    # Print file stats
    print('\nPer-file counts:')
    for f, c in stats.items():
        print(f'  {f}: {c}')


if __name__ == '__main__':
    main()
